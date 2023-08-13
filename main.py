import datetime
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.screenmanager import ScreenManager, Screen
from openpyxl import Workbook, load_workbook
import pandas as pd


EXCEL_FILE = 'time_tracker.xlsx'


class MainScreen(Screen):
    pass


class ResultsScreen(Screen):
    def __init__(self, **kwargs):
        super(ResultsScreen, self).__init__(**kwargs)
        self.result_label = Label(text='No Input has been provided yet')
        # the text of the label will be updated when the 'evaluate_data' function will run

        layout = BoxLayout(orientation='vertical')
        layout.add_widget(self.result_label)

        button_layout = BoxLayout(orientation='horizontal', size_hint=(1, 0.1))

        back_button = Button(text='Back', on_press=self.go_back)
        button_layout.add_widget(back_button)

        open_excel_button = Button(text='Open Excel Sheet', on_press=self.open_the_excel_sheet)
        button_layout.add_widget(open_excel_button)

        layout.add_widget(button_layout)

        self.add_widget(layout)

    def go_back(self, instance):
        self.manager.current = 'main'

    def open_the_excel_sheet(self, instance):
        self.manager.current = 'view_excel_screen'


class ExcelScreen(Screen):
    def __init__(self, **kwargs):
        super(ExcelScreen, self).__init__(**kwargs)
        layout = BoxLayout(orientation='vertical')
        self.excel_data_label = Label(text='')
        layout.add_widget(self.excel_data_label)

        button_layout = BoxLayout(orientation='horizontal', size_hint=(1, 0.1))
        back_to_main_button = Button(text='Back to INPUTs', on_press=self.go_back_to_main)
        back_to_results_button = Button(text='Back to RESULTs', on_press=self.go_back_to_results)
        button_layout.add_widget(back_to_main_button)
        button_layout.add_widget(back_to_results_button)

        try:
            data_stored = pd.read_excel(EXCEL_FILE,engine='openpyxl').to_string(col_space=int(self.width*0.25))
        except FileNotFoundError:
            workbook = Workbook()
            workbook.save(filename=EXCEL_FILE)
            data_stored = pd.read_excel(EXCEL_FILE,engine='openpyxl').to_string(col_space=int(self.width * 0.25))

        self.excel_data_label.text = data_stored

        layout.add_widget(button_layout)
        self.add_widget(layout)

    def go_back_to_main(self, instance):
        self.manager.current = 'main'

    def go_back_to_results(self, instance):
        self.manager.current = 'results'


class TimeTrackerApp(App):
    def build(self):
        self.screen_manager = ScreenManager()

        self.main_screen = MainScreen(name='main')
        self.results_screen = ResultsScreen(name='results')
        self.excel_screen = ExcelScreen(name='view_excel_screen')

        layout = BoxLayout(orientation='vertical')

        self.activity_started_label = Label(text='Activity started at:')
        self.initial_hour_input = TextInput(hint_text='Hour (1-12):', input_type='number', multiline=False)
        self.initial_hour_input.bind(on_text_validate=self.focus_initial_minute_input)

        self.initial_minute_input = TextInput(hint_text='Minutes (0-59):', input_type='number', multiline=False)
        self.initial_minute_input.bind(on_text_validate=self.focus_initial_time_period)

        self.initial_time_period_button_am = Button(text='AM', on_press=self.select_initial_period)
        self.initial_time_period_button_pm = Button(text='PM', on_press=self.select_initial_period)

        self.activity_ended_label = Label(text='Activity ended at:')
        self.final_hour_input = TextInput(hint_text='Hour (1-12):', input_type='number', multiline=False)
        self.final_hour_input.bind(on_text_validate=self.focus_final_minute_input)

        self.final_minute_input = TextInput(hint_text='Minutes (0-59):', input_type='number', multiline=False)
        self.final_minute_input.bind(on_text_validate=self.focus_final_time_period)

        self.final_time_period_button_am = Button(text='AM', on_press=self.select_final_period)
        self.final_time_period_button_pm = Button(text='PM', on_press=self.select_final_period)

        self.activity_input = TextInput(hint_text='Activity')
        self.submit_button = Button(text='Submit', on_press=self.submit_data)

        # Create the "Evaluate Today" button
        self.evaluate_today_button = Button(text='Evaluate Today', on_press=self.evaluate_today_data)

        # Create the "Evaluate Past Week" button
        self.evaluate_past_week_button = Button(text='Evaluate Past Week', on_press=self.evaluate_past_week_data)

        layout.add_widget(self.activity_started_label)
        layout.add_widget(self.initial_hour_input)
        layout.add_widget(self.initial_minute_input)
        layout.add_widget(self.initial_time_period_button_am)
        layout.add_widget(self.initial_time_period_button_pm)
        layout.add_widget(self.activity_ended_label)
        layout.add_widget(self.final_hour_input)
        layout.add_widget(self.final_minute_input)
        layout.add_widget(self.final_time_period_button_am)
        layout.add_widget(self.final_time_period_button_pm)
        layout.add_widget(Label(text='Activity:'))
        layout.add_widget(self.activity_input)
        layout.add_widget(self.submit_button)

        # Add the "Evaluate Today" button
        layout.add_widget(self.evaluate_today_button)

        # Add the "Evaluate Past Week" button
        layout.add_widget(self.evaluate_past_week_button)

        self.main_screen.add_widget(layout)

        self.screen_manager.add_widget(self.main_screen)
        self.screen_manager.add_widget(self.results_screen)
        self.screen_manager.add_widget(self.excel_screen)

        return self.screen_manager

    def focus_initial_minute_input(self, instance):
        self.initial_minute_input.focus = True

    def focus_initial_time_period(self, instance):
        self.initial_time_period_button_am.focus = True

    def focus_final_minute_input(self, instance):
        self.final_minute_input.focus = True

    def focus_final_time_period(self, instance):
        self.final_time_period_button_am.focus = True

    def select_initial_period(self, instance):
        self.initial_time_period_button_am.background_color = (1, 1, 1, 1)
        self.initial_time_period_button_pm.background_color = (1, 1, 1, 1)
        instance.background_color = (0.2, 0.6, 1, 1)
        self.selected_initial_period = instance.text

    def select_final_period(self, instance):
        self.final_time_period_button_am.background_color = (1, 1, 1, 1)
        self.final_time_period_button_pm.background_color = (1, 1, 1, 1)
        instance.background_color = (0.2, 0.6, 1, 1)
        self.selected_final_period = instance.text

    def submit_data(self, instance):
        initial_hour = self.initial_hour_input.text
        initial_minute = self.initial_minute_input.text
        final_hour = self.final_hour_input.text
        final_minute = self.final_minute_input.text
        activity = self.activity_input.text.title()  # Convert to title case


        try: self.selected_initial_period
        except: self.selected_initial_period = 'PM'  # default is set to PM, to avoid AttributeError when input is not provided.

        try: self.selected_final_period
        except: self.selected_final_period = 'PM'


        # Validate the input time
        if not self.validate_time(initial_hour, initial_minute, self.select_initial_period):
            self.display_error('Invalid start time')
            return

        if not self.validate_time(final_hour, final_minute, self.selected_final_period):
            self.display_error('Invalid end time')
            return

        # Calculate the duration
        duration = self.calculate_duration(initial_hour, initial_minute, final_hour, final_minute, self.selected_initial_period, self.selected_final_period)
        if duration <= datetime.timedelta():
            self.display_error('Invalid duration')
            return

        # Add the data to the Excel file
        self.update_excel(initial_hour, initial_minute, final_hour, final_minute, activity)

        # Clear input fields
        self.initial_hour_input.text = ''
        self.initial_minute_input.text = ''
        self.final_hour_input.text = ''
        self.final_minute_input.text = ''
        self.activity_input.text = ''

    def validate_time(self, hour, minute, waka):
        if not hour.isdigit() or not minute.isdigit():
            return False

        hour = int(hour)
        minute = int(minute)

        if hour < 1 or hour > 12 or minute < 0 or minute > 59:
            return False

        return True

    def calculate_duration(self, initial_hour, initial_minute, final_hour, final_minute, initial_period, final_period):
        initial_time = datetime.datetime.strptime(f'{initial_hour}:{initial_minute} {initial_period}', '%I:%M %p')
        final_time = datetime.datetime.strptime(f'{final_hour}:{final_minute} {final_period}', '%I:%M %p')
        duration = final_time - initial_time
        return duration

    def display_error(self, error_message):
        self.main_screen.add_widget(Label(text=error_message))

    def update_excel(self, initial_hour, initial_minute, final_hour, final_minute, activity):
        # Create or open the Excel file
        try:
            workbook = load_workbook(filename=EXCEL_FILE)
        except FileNotFoundError:
            workbook = Workbook()
            workbook.save(filename=EXCEL_FILE)

        sheet = workbook.active

        # Check if the sheet has been populated previously
        if sheet['A1'].value is None:
            # Add headers if the sheet is empty
            sheet['A1'] = 'Date'
            sheet['B1'] = 'Initial Time'
            sheet['C1'] = 'Final Time'
            sheet['D1'] = 'Activity'

        # Find the next available row
        row_number = sheet.max_row + 1

        # Get the current date
        current_date = datetime.datetime.now().date()

        # Prepare the initial and final time
        initial_time = self.prepare_time_string(initial_hour, initial_minute, self.selected_initial_period)
        final_time = self.prepare_time_string(final_hour, final_minute, self.selected_final_period)

        # Add the data to the Excel file
        sheet.cell(row=row_number, column=1).value = current_date
        sheet.cell(row=row_number, column=2).value = initial_time
        sheet.cell(row=row_number, column=3).value = final_time
        sheet.cell(row=row_number, column=4).value = activity

        # Save the Excel file
        workbook.save(filename=EXCEL_FILE)

    def prepare_time_string(self, hour, minute, period):
        time_string = f'{hour}:{minute}'
        if period == 'PM':
            time_string += ' PM'
        else:
            time_string += ' AM'
        return time_string

    def evaluate_today_data(self, instance):
        self.evaluate_data(instance, [datetime.datetime.now().date()])

    def evaluate_past_week_data(self, instance):
        current_date = datetime.datetime.now().date()
        past_week = [current_date - datetime.timedelta(days=i) for i in range(7)]
        self.evaluate_data(instance, past_week)

    # ...

    def evaluate_data(self, instance, dates):
        self.results_screen.result_label.text = ''  # Clear previous results
        self.screen_manager.current = 'results'

        # Open the Excel file
        workbook = load_workbook(filename=EXCEL_FILE)
        sheet = workbook.active

        # Calculate the total hours and minutes for each activity
        activity_totals = {}
        for row in sheet.iter_rows(min_row=2):
            record_date = row[0].value
            if isinstance(record_date, datetime.datetime):
                record_date = record_date.date()
            if record_date in dates:
                activity = row[3].value
                initial_time = datetime.datetime.strptime(row[1].value, '%I:%M %p')
                final_time = datetime.datetime.strptime(row[2].value, '%I:%M %p')
                duration = final_time - initial_time

                if activity in activity_totals:
                    activity_totals[activity] += duration
                else:
                    activity_totals[activity] = duration

        # Display the results
        result_data = ''
        result_data += f'RESULTS\n'
        for activity, total_duration in activity_totals.items():
            total_hours = total_duration.total_seconds() // 3600
            total_minutes = (total_duration.total_seconds() % 3600) // 60

            result_data += f'{activity}: {int(total_hours)} hours, {int(total_minutes)} minutes\n'

        self.results_screen.result_label.text = result_data


if __name__ == '__main__':
    TimeTrackerApp().run()